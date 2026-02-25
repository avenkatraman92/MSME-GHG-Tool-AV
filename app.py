import pandas as pd
import streamlit as st
from io import BytesIO

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

st.set_page_config(page_title="MSME GHG Tool", layout="wide")

st.title("MSME GHG Data Collection Tool")
st.write(
    "This tool is designed for MSME data collection for Scope 1 and Scope 2 inventory preparation. "
    "It supports multi-unit businesses and creates a structured workbook for consulting-led GHG assessment."
)

# ─── 0. ORGANISATIONAL BOUNDARY ───────────────────────────────────────────────
st.header("0. Organisational Boundary")
col_b1, col_b2 = st.columns(2)
with col_b1:
    boundary_approach = st.selectbox(
        "Boundary Approach (GHG Protocol) *",
        ["Operational Control", "Financial Control", "Equity Share"],
    )
with col_b2:
    consolidation_note = st.text_input(
        "Consolidation Notes (optional)",
        placeholder="e.g. Joint ventures excluded, leased facilities included under operational control",
    )
boundary_exclusions = st.text_area(
    "Boundary Exclusions (optional)",
    placeholder="List any facilities, operations, or sources deliberately excluded from this inventory and the reason.",
    height=80,
)

# ─── 1. COMPANY PROFILE ───────────────────────────────────────────────────────
st.header("1. Company Profile")
col1, col2 = st.columns(2)

with col1:
    company_name = st.text_input("Company Name *")
    udyam_id = st.text_input("UDYAM ID")
    industry = st.text_input("Industry *")
    nic_code = st.text_input("NIC Code (India)")

with col2:
    isic_code = st.text_input("International Industry Code (ISIC)")
    msme_class = st.selectbox("MSME Classification", ["Micro", "Small", "Medium"])
    turnover = st.number_input("Annual Turnover (INR)", min_value=0.0, step=100000.0)

    # Fix 2: Extended reporting year range + custom option
    year_options = [f"{y}-{str(y + 1)[-2:]}" for y in range(2018, 2030)] + ["Custom"]
    reporting_year_select = st.selectbox("Reporting Year *", year_options, index=5)
    if reporting_year_select == "Custom":
        reporting_year = st.text_input("Enter Reporting Year (e.g. 2026-27)")
    else:
        reporting_year = reporting_year_select

st.subheader("Employees")
col3, col4 = st.columns(2)
with col3:
    perm_m = st.number_input("Permanent Employees - Male", min_value=0, key="perm_m")
    perm_f = st.number_input("Permanent Employees - Female", min_value=0, key="perm_f")
with col4:
    cont_m = st.number_input("Contract Employees - Male", min_value=0, key="cont_m")
    cont_f = st.number_input("Contract Employees - Female", min_value=0, key="cont_f")

# ─── 2. UNITS SETUP ───────────────────────────────────────────────────────────
st.header("2. Units Setup")
number_of_units = st.number_input(
    "How many operational units/facilities should be covered?",
    min_value=1, max_value=20, value=1, step=1,
)

UNIT_SOURCE_CATALOG = {
    "DG Set": ("Stationary Combustion", "Diesel Consumption", "Litres", "Fuel purchase bills", "Scope 1"),
    "Boiler": ("Stationary Combustion", "Fuel Consumption", "kg / Litres / SCM", "Fuel purchase and logbooks", "Scope 1"),
    "Furnace": ("Stationary Combustion", "Fuel Consumption", "kg / Litres / SCM", "Fuel purchase and process logs", "Scope 1"),
    "Thermic Fluid Heater": ("Stationary Combustion", "Fuel Consumption", "kg / Litres / SCM", "Fuel purchase and maintenance logs", "Scope 1"),
    "Kiln / Oven": ("Stationary Combustion", "Fuel Consumption", "kg / Litres / SCM", "Fuel purchase and production logs", "Scope 1"),
    "Process Steam Generation": ("Stationary Combustion", "Fuel Consumption", "kg / Litres / SCM", "Steam generation and fuel records", "Scope 1"),
    "Industrial HVAC": ("Fugitive Emissions", "Refrigerant Refill / Leakage", "kg/year", "Service and maintenance records", "Scope 1"),
    "Air Conditioners": ("Fugitive Emissions", "Refrigerant Refill / Leakage", "kg/year", "AMC/service records", "Scope 1"),
    "Chillers": ("Fugitive Emissions", "Refrigerant Refill / Leakage", "kg/year", "Service and maintenance records", "Scope 1"),
    "Cold Storage / Refrigeration": ("Fugitive Emissions", "Refrigerant Refill / Leakage", "kg/year", "Maintenance and gas refill records", "Scope 1"),
    "Fire Suppression Systems": ("Fugitive Emissions", "Refill Quantity", "kg/year", "Inspection and refill records", "Scope 1"),
    "Welding with CO2 shielding gas": ("Process / Fugitive", "CO2 Consumption", "kg/month", "Purchase invoices and usage logs", "Scope 1"),
    "LPG/CNG used in process": ("Stationary Combustion", "Fuel Consumption", "kg / SCM", "Fuel purchase bills", "Scope 1"),
    "Electricity Consumption": ("Purchased Electricity", "Electricity Consumption", "kWh", "Electricity bills", "Scope 2"),
}

VEHICLE_SOURCE_CATALOG = {
    "Two-wheelers (Petrol)": ("Mobile Combustion", "Petrol Consumption", "Litres", "Fuel logs / bills", "Scope 1"),
    "Cars (Petrol)": ("Mobile Combustion", "Petrol Consumption", "Litres", "Fuel logs / bills", "Scope 1"),
    "Cars (Diesel)": ("Mobile Combustion", "Diesel Consumption", "Litres", "Fuel logs / bills", "Scope 1"),
    "Light Commercial Vehicles": ("Mobile Combustion", "Diesel/CNG Consumption", "Litres / kg", "Fuel logs / bills", "Scope 1"),
    "Heavy Commercial Vehicles": ("Mobile Combustion", "Diesel Consumption", "Litres", "Fuel logs / bills", "Scope 1"),
    "Forklifts / Material Handling": ("Mobile Combustion", "Fuel/Energy Consumption", "Litres / kWh", "Fuel logs / charging logs", "Scope 1"),
}

HO_SOURCE_CATALOG = {
    "Head Office Electricity": ("Purchased Electricity", "Electricity Consumption", "kWh", "Electricity bills", "Scope 2"),
    "Head Office AC / Refrigerants": ("Fugitive Emissions", "Refrigerant Refill / Leakage", "kg/year", "AMC/service records", "Scope 1"),
    "Head Office DG Set": ("Stationary Combustion", "Diesel Consumption", "Litres", "Fuel purchase records", "Scope 1"),
}

ALL_FUELS = ["Diesel", "LPG", "Natural Gas", "Coal", "Biomass", "FO/HSD", "Other"]

units_data = []
for i in range(int(number_of_units)):
    unit_no = i + 1
    with st.expander(f"Unit {unit_no} Profile and Module", expanded=(i == 0)):
        c1, c2 = st.columns(2)
        with c1:
            unit_name = st.text_input("Unit Name", value=f"Unit {unit_no}", key=f"unit_name_{i}")
            unit_location = st.text_input("Location (City/State)", key=f"unit_location_{i}")
            unit_area = st.number_input("Area (sq ft)", min_value=0.0, step=100.0, key=f"unit_area_{i}")
        with c2:
            unit_ownership = st.selectbox(
                "Ownership",
                ["Fully Owned", "Partly Owned", "Leased", "Contract Manufacturing"],
                key=f"unit_ownership_{i}",
            )
            principal_process = st.text_input("Primary production process", key=f"unit_process_{i}")
            annual_production = st.text_input("Annual production / throughput (optional)", key=f"unit_production_{i}")

        st.caption(
            "Unit module covers stationary combustion, fugitive emissions, and purchased electricity. "
            "Do not include vehicle fuel use here — vehicles are captured in the Overall Module."
        )

        # Fix 4: Select all checkbox for unit sources
        all_unit_keys = list(UNIT_SOURCE_CATALOG.keys())
        select_all_unit = st.checkbox("Select all emission sources for this unit", key=f"sel_all_unit_{i}")
        selected_sources = st.multiselect(
            "Select machinery / process sources applicable for this unit",
            all_unit_keys,
            default=all_unit_keys if select_all_unit else [],
            key=f"unit_sources_{i}",
        )

        include_general_fuel = st.checkbox(
            "Include general stationary fuel use (exclude vehicle fuels)",
            key=f"include_general_fuel_{i}",
        )
        general_fuels = []
        if include_general_fuel:
            # Fix 4: Select all for fuel types
            select_all_fuels = st.checkbox("Select all fuel types", key=f"sel_all_fuels_{i}")
            general_fuels = st.multiselect(
                "Select general stationary fuels",
                ALL_FUELS,
                default=ALL_FUELS if select_all_fuels else [],
                key=f"general_fuels_{i}",
            )

        units_data.append({
            "unit_name": unit_name,
            "location": unit_location,
            "area_sqft": unit_area,
            "ownership": unit_ownership,
            "primary_process": principal_process,
            "annual_production": annual_production,
            "selected_sources": selected_sources,
            "general_fuels": general_fuels,
        })

# ─── 3. OVERALL MODULE ────────────────────────────────────────────────────────
st.header("3. Overall Module (Common / Shared Sources)")
st.caption("Capture mobile combustion (vehicles) and head-office emissions not already included at unit level.")

# Fix 4: Select all for vehicles and HO
all_vehicle_keys = list(VEHICLE_SOURCE_CATALOG.keys())
sel_all_veh = st.checkbox("Select all vehicle sources")
vehicle_sources = st.multiselect(
    "Vehicle emissions sources (common/shared across units)",
    all_vehicle_keys,
    default=all_vehicle_keys if sel_all_veh else [],
)

all_ho_keys = list(HO_SOURCE_CATALOG.keys())
sel_all_ho = st.checkbox("Select all head office / common sources")
ho_sources = st.multiselect(
    "Head office / common electricity and fugitive sources",
    all_ho_keys,
    default=all_ho_keys if sel_all_ho else [],
)

# ─── 4. GENERATE ──────────────────────────────────────────────────────────────
st.header("4. Download Data Collection Template")

# ── Column layout (Fix 6: Scope moved to column 2; Fix 11: Data Status added) ─
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
COLUMNS = (
    ["Category", "Scope", "Emission Source", "Equipment / Activity", "Data Required", "Unit"]
    + MONTHS
    + ["Annual Total", "Data Status", "Guidance"]
)

# 1-indexed column positions for openpyxl
COL_CATEGORY = 1
COL_SCOPE    = 2
COL_SOURCE   = 3
COL_EQUIP    = 4
COL_DATAREQ  = 5
COL_UNIT     = 6
COL_JAN      = 7
COL_DEC      = 18
COL_ANNUAL   = 19
COL_STATUS   = 20
COL_GUIDANCE = 21
TOTAL_COLS   = 21


def build_rows_from_catalog(selected_items, catalog):
    rows = []
    for item in selected_items:
        category, data_required, unit, guidance, scope = catalog[item]
        rows.append(
            [category, scope, item, item, data_required, unit]
            + [""] * 12
            + ["", "", guidance]
        )
    return rows


def unit_sheet_dataframe(unit):
    rows = []
    rows.append(
        ["UNIT GUIDANCE", "", "", "", "", ""] + [""] * 12
        + ["", "", "Exclude vehicle fuel use in this sheet."]
    )
    rows.extend(build_rows_from_catalog(unit["selected_sources"], UNIT_SOURCE_CATALOG))

    # Fix 3: each fuel gets its own clearly named row
    for fuel in unit["general_fuels"]:
        rows.append(
            ["Stationary Combustion", "Scope 1", f"General Fuel — {fuel}", fuel,
             "Fuel Consumption", "Litres / kg / SCM"]
            + [""] * 12
            + ["", "", "Non-vehicle stationary/process fuel only"]
        )

    if len(rows) <= 1:
        rows.append(["No sources selected", "", "", "", "", ""] + [""] * 12 + ["", "", ""])

    return pd.DataFrame(rows, columns=COLUMNS)


def overall_sheet_dataframe(vehicle_selection, ho_selection):
    rows = []
    rows.append(
        ["OVERALL GUIDANCE", "", "", "", "", ""] + [""] * 12
        + ["", "", "Capture shared/common sources only. Do not duplicate unit-level data."]
    )
    rows.extend(build_rows_from_catalog(vehicle_selection, VEHICLE_SOURCE_CATALOG))
    rows.extend(build_rows_from_catalog(ho_selection, HO_SOURCE_CATALOG))

    if len(rows) == 1:
        rows.append(["No shared sources selected", "", "", "", "", ""] + [""] * 12 + ["", "", ""])

    return pd.DataFrame(rows, columns=COLUMNS)


# ── Styling helpers ────────────────────────────────────────────────────────────

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL   = PatternFill("solid", fgColor="305496")
HEADER_FONT   = Font(color="FFFFFF", bold=True)
SCOPE1_FILL   = PatternFill("solid", fgColor="E2EFDA")   # light green
SCOPE2_FILL   = PatternFill("solid", fgColor="DDEEFF")   # light blue
GUIDANCE_FILL = PatternFill("solid", fgColor="D9D9D9")
YELLOW_FILL   = PatternFill("solid", fgColor="FFF2CC")
STATUS_FILL   = PatternFill("solid", fgColor="FCE4D6")   # light salmon
LABEL_FILL    = PatternFill("solid", fgColor="E7E6E6")
BOLD_FONT     = Font(bold=True)


def style_data_sheet(ws, data_start_row):
    """Apply styling to a unit or overall data sheet."""
    max_row = ws.max_row
    header_row = data_start_row - 1

    # Header row
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    # Data rows
    for r in range(data_start_row, max_row + 1):
        row_val = str(ws.cell(r, 1).value or "")
        is_guidance_row = "GUIDANCE" in row_val.upper()
        is_empty_row = row_val in ("No sources selected", "No shared sources selected", "")

        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(r, c)

            if is_guidance_row:
                cell.fill = GUIDANCE_FILL
                cell.font = BOLD_FONT
            elif c == COL_SCOPE:
                sv = str(cell.value or "")
                if sv == "Scope 1":
                    cell.fill = SCOPE1_FILL
                elif sv == "Scope 2":
                    cell.fill = SCOPE2_FILL
            elif COL_JAN <= c <= COL_DEC:
                cell.fill = YELLOW_FILL
                cell.border = THIN_BORDER
            elif c == COL_ANNUAL:
                cell.border = THIN_BORDER
                cell.font = BOLD_FONT
            elif c == COL_STATUS:
                cell.fill = STATUS_FILL

            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in (COL_GUIDANCE, COL_SOURCE, COL_EQUIP)),
            )

        # Fix 7: annual SUM formula for data rows
        if not is_guidance_row and not is_empty_row:
            jan = f"{get_column_letter(COL_JAN)}{r}"
            dec = f"{get_column_letter(COL_DEC)}{r}"
            ws.cell(r, COL_ANNUAL).value = f"=SUM({jan}:{dec})"

    # Fix 11: dropdown data validation for Data Status
    dv = DataValidation(
        type="list",
        formula1='"Metered,Estimated,Partially Available,Not Available"',
        allow_blank=True,
        showDropDown=False,
    )
    status_range = (
        f"{get_column_letter(COL_STATUS)}{data_start_row}"
        f":{get_column_letter(COL_STATUS)}{max_row}"
    )
    dv.sqref = status_range
    ws.add_data_validation(dv)

    # Fix 9: explicit column widths — no arbitrary cap on guidance
    col_widths = {
        COL_CATEGORY: 22, COL_SCOPE: 11, COL_SOURCE: 28, COL_EQUIP: 24,
        COL_DATAREQ: 20, COL_UNIT: 14, COL_ANNUAL: 14,
        COL_STATUS: 20, COL_GUIDANCE: 52,
    }
    for c in range(1, TOTAL_COLS + 1):
        ltr = get_column_letter(c)
        if c in col_widths:
            ws.column_dimensions[ltr].width = col_widths[c]
        elif COL_JAN <= c <= COL_DEC:
            ws.column_dimensions[ltr].width = 8

    ws.freeze_panes = f"{get_column_letter(COL_JAN)}{data_start_row}"


def style_unit_summary(ws, unit):
    """Fix 5: styled summary block at top of each unit sheet (rows 1–6)."""
    SUMMARY_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    SUMMARY_LBL_FILL = PatternFill("solid", fgColor="BDD7EE")
    SUMMARY_VAL_FILL = PatternFill("solid", fgColor="DDEBF7")
    WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)
    LBL_FONT = Font(bold=True, size=9)
    VAL_FONT = Font(size=9)

    # Row 1: merged header
    ws.merge_cells("A1:U1")
    c = ws["A1"]
    c.value = "UNIT SUMMARY"
    c.fill = SUMMARY_HDR_FILL
    c.font = WHITE_BOLD
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Rows 2–4: label / value pairs
    pairs = [
        ("A2", "Unit Name",          "B2", unit["unit_name"]),
        ("C2", "Ownership",          "D2", unit["ownership"]),
        ("A3", "Location",           "B3", unit["location"]),
        ("C3", "Primary Process",    "D3", unit["primary_process"]),
        ("A4", "Area (sq ft)",       "B4", unit["area_sqft"]),
        ("C4", "Annual Production",  "D4", unit["annual_production"]),
    ]
    for lbl_ref, lbl_val, val_ref, val_val in pairs:
        lc = ws[lbl_ref]
        lc.value = lbl_val
        lc.fill = SUMMARY_LBL_FILL
        lc.font = LBL_FONT
        lc.border = THIN_BORDER
        lc.alignment = Alignment(horizontal="left", vertical="center")

        vc = ws[val_ref]
        vc.value = val_val
        vc.fill = SUMMARY_VAL_FILL
        vc.font = VAL_FONT
        vc.border = THIN_BORDER
        vc.alignment = Alignment(horizontal="left", vertical="center")

    # Row 5: blank spacer; Row 6: guidance note
    ws.merge_cells("A6:U6")
    g = ws["A6"]
    g.value = (
        "⚠  Guidance: Capture only unit-level stationary combustion, fugitive emissions, and "
        "purchased electricity. Do not include vehicle fuel use — vehicles go in the Overall Module."
    )
    g.fill = PatternFill("solid", fgColor="FFF2CC")
    g.font = Font(italic=True, size=9)
    g.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[6].height = 28


def style_about_sheet(ws):
    """Style the About & Guidance sheet."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).fill = LABEL_FILL
        ws.cell(r, 1).font = BOLD_FONT
        ws.cell(r, 1).alignment = Alignment(vertical="top")
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 90
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"


def style_profile_sheet(ws):
    """Fix 8 + Fix 12: style the Company Profile sheet."""
    SECTION_FILL = PatternFill("solid", fgColor="1F4E79")
    SECTION_FONT = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or "")
        # Section dividers (all-caps labels like "COMPANY PROFILE")
        if val.isupper() and val.strip():
            ws.merge_cells(f"A{r}:B{r}")
            ws.cell(r, 1).fill = SECTION_FILL
            ws.cell(r, 1).font = SECTION_FONT
            ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 18
        else:
            ws.cell(r, 1).fill = LABEL_FILL
            ws.cell(r, 1).font = BOLD_FONT
            ws.cell(r, 1).alignment = Alignment(vertical="center")
            ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[r].height = 16

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 65


def add_branding_footer(ws):
    mr = ws.max_row
    ws.cell(mr + 2, 1).value = "Developed by Arun Venkatraman | Total Impact"
    ws.cell(mr + 3, 1).value = "Sustainability & ESG Advisory | GHG Accounting | Supply Chain Decarbonisation"
    ws.cell(mr + 4, 1).value = "arun@totalimpact.co.in | +91 6374350144"
    for r in [mr + 2, mr + 3, mr + 4]:
        ws.cell(r, 1).font = Font(italic=True, color="595959", size=8)


# ── Excel generation ───────────────────────────────────────────────────────────

def generate_excel(units, vehicle_selection, ho_selection):
    output = BytesIO()

    about_rows = [
        ["ABOUT THE TOOL",
         "This MSME greenhouse gas data collection workbook is a practical bridge between operational "
         "records and a complete emissions inventory. It supports structured data gathering across multiple "
         "units, enables consistent monthly evidence capture, and improves readiness for consulting-led GHG "
         "accounting and reporting. The workbook focuses on activity data and documentation quality — because "
         "accurate emissions estimation depends on complete source identification, correct units, and "
         "verifiable source records."],
        ["GUIDANCE FOR USE",
         "Start with the Company Profile sheet and confirm reporting-year boundaries before filling source "
         "sheets. For each unit sheet, include only unit-level stationary combustion, process/fugitive "
         "emissions, and purchased electricity. Do not enter vehicle fuel use within unit sheets — mobile "
         "combustion is captured in the Overall Module to avoid double-counting. Record values month-wise "
         "using invoices, utility bills, refill records, and maintenance logs. Leave cells blank where data "
         "is unavailable and mark the Data Status column accordingly instead of estimating."],
        ["DATA STATUS COLUMN",
         "Use the dropdown in the 'Data Status' column for every row:\n"
         "  • Metered — directly read from a meter, sub-meter, or utility bill\n"
         "  • Estimated — calculated or approximated from indirect evidence\n"
         "  • Partially Available — some months have data, others do not\n"
         "  • Not Available — no data exists; source will need follow-up"],
        ["KEY SUGGESTIONS",
         "— Map metering boundaries before data entry\n"
         "— Confirm whether head office electricity is already included in a unit sheet\n"
         "— Track refrigerant type and refill date together in the same row\n"
         "— Separate process fuel from transport fuel\n"
         "— Preserve invoice copies with clear period labels\n"
         "— Align monthly data cut-off dates with your financial closure cycle\n"
         "— The Annual Total column uses a SUM formula — do not overwrite it"],
        ["CONTACT",
         "Developed by Arun Venkatraman | Total Impact | arun@totalimpact.co.in | +91 6374350144"],
    ]
    about_df = pd.DataFrame(about_rows, columns=["Section", "Details"])

    # Fix 8 + Fix 12: separate Company Profile sheet including boundary
    profile_rows = [
        ["COMPANY PROFILE", ""],
        ["Company Name", company_name],
        ["UDYAM ID", udyam_id],
        ["Industry", industry],
        ["NIC Code (India)", nic_code],
        ["International Code (ISIC)", isic_code],
        ["MSME Classification", msme_class],
        ["Annual Turnover (INR)", turnover],
        ["Reporting Year", reporting_year],
        ["Permanent Employees — Male", perm_m],
        ["Permanent Employees — Female", perm_f],
        ["Contract Employees — Male", cont_m],
        ["Contract Employees — Female", cont_f],
        ["Number of Units / Facilities", len(units)],
        ["", ""],
        ["ORGANISATIONAL BOUNDARY", ""],
        ["Boundary Approach (GHG Protocol)", boundary_approach],
        ["Consolidation Notes", consolidation_note],
        ["Boundary Exclusions", boundary_exclusions],
    ]
    profile_df = pd.DataFrame(profile_rows, columns=["Field", "Value"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # Sheet 1: About & Guidance
        about_df.to_excel(writer, sheet_name="About & Guidance", index=False)

        # Sheet 2: Company Profile (Fix 8: split from About)
        profile_df.to_excel(writer, sheet_name="Company Profile", index=False)

        # Unit sheets
        for idx, unit in enumerate(units, start=1):
            unit_df = unit_sheet_dataframe(unit)
            sname = f"Unit {idx} Module"
            unit_df.to_excel(writer, sheet_name=sname, index=False, startrow=6)
            ws = writer.sheets[sname]
            style_unit_summary(ws, unit)          # Fix 5
            style_data_sheet(ws, data_start_row=8)

        # Overall Module
        overall_df = overall_sheet_dataframe(vehicle_selection, ho_selection)
        overall_df.to_excel(writer, sheet_name="Overall Module", index=False)
        style_data_sheet(writer.sheets["Overall Module"], data_start_row=2)

        # Style About & Guidance
        style_about_sheet(writer.sheets["About & Guidance"])

        # Style Company Profile
        style_profile_sheet(writer.sheets["Company Profile"])

        # Branding footer on all sheets
        for ws in writer.sheets.values():
            add_branding_footer(ws)

    output.seek(0)
    return output


# Fix 1: validate required fields before generating
if st.button("Generate Excel Template"):
    errors = []
    if not company_name.strip():
        errors.append("Company Name is required.")
    if not industry.strip():
        errors.append("Industry is required.")
    if reporting_year_select == "Custom" and not reporting_year.strip():
        errors.append("Please enter a custom reporting year.")

    if errors:
        for e in errors:
            st.error(e)
    else:
        with st.spinner("Building your Excel template…"):
            excel_file = generate_excel(units_data, vehicle_sources, ho_sources)

        safe_name = company_name.strip().replace(" ", "_")
        st.download_button(
            label="⬇  Download Excel Template",
            data=excel_file,
            file_name=f"MSME_GHG_Template_{safe_name}_{reporting_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.success("Template ready. Click the button above to download.")

st.markdown("---")
st.markdown(
    """
    **Developed by Arun Venkatraman | Total Impact**
    Sustainability & ESG Advisory | GHG Accounting | Supply Chain Decarbonisation
    📧 arun@totalimpact.co.in | 📞 +91 6374350144
    """
)
